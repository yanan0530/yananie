from django.shortcuts import render, redirect
from django.http import HttpResponse

from rest_framework import serializers
from rest_framework.response import Response
from rest_framework.decorators import api_view

from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate

# from .models import UserInfo

import openpyxl as xl


# class UserInfoJson(serializers.ModelSerializer):
#     class Meta:
#         depth = 1
#         model = UserInfo
#         fields = '__all__'


@api_view(['GET'])
def index(request):
    res = 1
    if not isinstance(request.user, User):
        res = 0
        return Response({'res': res})
    user = request.user
    # userinfo = UserInfo.objects.get(userinfo=user)
    # json1 = UserInfoJson(userinfo)
    context = {'res': res, }
    return Response(context)


@api_view(['POST'])
def loginCms(request):
    username = request.data['username']
    password = request.data['password']
    user = authenticate(username=username, password=password)
    if user != None:
        login(request, user)
        return Response({'res': 1})
    else:
        return Response({'res': 0})


@api_view(['POST'])
def fileExcel(request):
    file = request.FILES.get('file')
    wb = xl.load_workbook(file)
    ws = wb.active
    datas = []
    columns = []
    for x in range(1, 2):
        for y in range(1, ws.max_column + 1):
            key = xl.utils.cell.get_column_letter(y)
            column = {}
            column['title'] = ws.cell(row=x, column=y).value
            column['key'] = key
            columns.append(column)
    for x in range(3, 8):
        data = {}
        for y in range(1, ws.max_column + 1):
            key = xl.utils.cell.get_column_letter(y)
            data[key] = ws.cell(row=x, column=y).value
        datas.append(data)
    user = request.user
    url = f'upload/{user}.xlsx'
    wb.save(str(url))
    return Response({"res": 0, 'url': url, "columns": columns, 'datas': datas})


@api_view(['POST'])
def splicingCell(request):

    file = request.FILES.get('file')
    cell = request.data['cell']
    character = request.data['character']

    wb = xl.load_workbook(file)
    ws = wb.active

    str = ""
    for i in range(2, ws.max_row + 1):
        position=f'{cell}{i}'
        c = ws[position].value
        str += f"{c}{character}"
    context = {
        "str": str[:-1]
    }
    return Response(context)
